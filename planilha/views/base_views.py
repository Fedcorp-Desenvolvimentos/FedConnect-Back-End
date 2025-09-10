# meu_app_planilhas/views/base_planilha.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def _base_download_model_excel(sheet_title, header_name, example_values, column_width, filename):
    """
    Função base para gerar e baixar um arquivo Excel modelo.
    Pode ser reutilizada por diferentes views de download de modelo.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    # Adicionar cabeçalho
    sheet.cell(row=1, column=1, value=header_name)
    
    # Estiliza o cabeçalho
    header_cell = sheet.cell(row=1, column=1)
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_cell.border = thin_border
    sheet.column_dimensions[get_column_letter(1)].width = column_width

    # Adicionar valores de exemplo
    for row_num, value in enumerate(example_values, 2):
        sheet.cell(row=row_num, column=1, value=value)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _base_format_error_result(entrada_original, error_message, result_keys):
    """
    Formata um resultado de erro padrão para as linhas da planilha de saída.
    Preenche campos de dados com vazio e adiciona status de falha e detalhes do erro.
    `result_keys` são as chaves esperadas nos resultados de sucesso (excluindo STATUS/ERRO).
    """
    formatted = {key: "" for key in result_keys}
    formatted["STATUS CONSULTA"] = "FALHA"
    formatted["ERRO_DETALHES"] = error_message
    
    # Adiciona a entrada original no campo correto, dependendo do tipo (CNPJ, CPF, etc.)
    # Isso exige que a chave "CNPJ (Entrada)" ou "CPF (Entrada)" etc. exista nas colunas.
    if "CNPJ (Entrada)" in formatted:
        formatted["CNPJ (Entrada)"] = entrada_original
    elif "CPF (Entrada)" in formatted:
        formatted["CPF (Entrada)"] = entrada_original
    elif "CEP (Entrada)" in formatted: # Adicionado para CEP, se aplicável
        formatted["CEP (Entrada)"] = entrada_original
    else: # Fallback genérico se a chave de entrada não for específica
        formatted["Entrada Original"] = entrada_original # Certifique-se de que esta chave esteja no seu column_defs se for usada.

    return formatted