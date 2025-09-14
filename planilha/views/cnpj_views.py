from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.authentication import JWTAuthentication


import requests
import uuid
import logging

from ..serializers import ProcessamentoPlanilhaCnpjInputSerializer
from .base_views import _base_download_model_excel, _base_format_error_result # Assumindo que essas funções são eficientes

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.conf import settings 

logger = logging.getLogger(__name__)

# --- Recomendo fortemente que esta URL não seja fixa no código ---
# Ela DEVE vir de uma variável de ambiente em produção.
YOUR_CONSULTA_API_URL = settings.CONSULTA_API_URL

@api_view(["GET"])
@permission_classes([AllowAny])
def baixar_planilha_modelo_drf_cnpj(request):
    """
    View para gerar e baixar a planilha modelo de consulta de CNPJs.
    """
    return _base_download_model_excel(
        sheet_title="CNPJs para Consulta",
        header_name="CNPJ",
        example_values=["33.647.553/0001-90", "33649575000199", "41.096.250/0001-54"],
        column_width=25,
        filename="planilha-modelo-cnpj.xlsx",
    )


class ProcessarPlanilhaCnpjsView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated]

    cnpj_column_defs = [
        {"header": "CNPJ (Entrada)", "key": "CNPJ (Entrada)", "width": 25},
        {"header": "STATUS CONSULTA", "key": "STATUS CONSULTA", "width": 18},
        {"header": "Razão Social", "key": "Razão Social", "width": 40},
        {"header": "CNPJ (Retornado)", "key": "CNPJ", "width": 25},
        {"header": "Atividade Principal", "key": "Atividade Principal", "width": 50},
        {"header": "Telefone", "key": "Telefone", "width": 15},
        {"header": "Telefone 2", "key": "Telefone 2", "width": 15},
        {"header": "Logradouro", "key": "Logradouro", "width": 40},
        {"header": "Número", "key": "Número", "width": 15},
        {"header": "Complemento", "key": "Complemento", "width": 20},
        {"header": "Bairro", "key": "Bairro", "width": 25},
        {"header": "Município", "key": "Município", "width": 25},
        {"header": "UF", "key": "UF", "width": 10},
        {"header": "CEP", "key": "CEP", "width": 15},
        {"header": "Situação Cadastral", "key": "Situação Cadastral", "width": 25},
        {"header": "ERRO_DETALHES", "key": "ERRO_DETALHES", "width": 40},
    ]

    cnpj_result_keys = [
        col["key"]
        for col in cnpj_column_defs
        if col["key"] not in ["STATUS CONSULTA", "ERRO_DETALHES"]
    ]
    
    # --- Métodos de formatação para clareza ---
    def _format_success_result(self, cnpj_entrada, api_cnpj_data):
        """
        Formata os dados de sucesso para a linha da planilha de CNPJ.
        """
        return {
            "CNPJ (Entrada)": cnpj_entrada,
            "STATUS CONSULTA": "SUCESSO",
            "ERRO_DETALHES": "",
            "Razão Social": api_cnpj_data.get("razao_social", ""),
            "CNPJ": api_cnpj_data.get("cnpj", ""),
            "Atividade Principal": api_cnpj_data.get("cnae_fiscal_descricao", ""),
            "Telefone": api_cnpj_data.get("ddd_telefone_1", ""),
            "Telefone 2": api_cnpj_data.get("ddd_telefone_2", ""),
            "Logradouro": api_cnpj_data.get("logradouro", ""),
            "Número": api_cnpj_data.get("numero", ""),
            "Complemento": api_cnpj_data.get("complemento", ""),
            "Bairro": api_cnpj_data.get("bairro", ""),
            "Município": api_cnpj_data.get("municipio", ""),
            "UF": api_cnpj_data.get("uf", ""),
            "CEP": api_cnpj_data.get("cep", ""),
            "Situação Cadastral": api_cnpj_data.get("descricao_situacao_cadastral", ""),
        }

    def _format_error_result(self, cnpj_entrada, error_message):
        """
        Formata os dados de erro para a linha da planilha de CNPJ, utilizando o helper base.
        """
        return _base_format_error_result(cnpj_entrada, error_message, self.cnpj_result_keys)


    def post(self, request, *args, **kwargs):
        logger.info("Iniciando ProcessarPlanilhaCnpjsView.post")
        if not YOUR_CONSULTA_API_URL:
            logger.error("URL da API de consulta não configurada.")
            return Response(
                {"message": "URL da API de consulta não configurada no backend."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        serializer = ProcessamentoPlanilhaCnpjInputSerializer(data=request.data)
        if not serializer.is_valid():
            logger.warning(f"Dados do serializer inválidos: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        valid_data = serializer.validated_data
        cnpjs_from_excel = valid_data.get("cnpjs", [])
        origem_consulta = valid_data.get("origem", "planilha")

        batch_id_para_planilha = None
        if origem_consulta == "planilha":
            batch_id_para_planilha = uuid.uuid4()
            
        user_jwt_token = None
        if request.auth:
            user_jwt_token = str(request.auth) 
        if not user_jwt_token:
            logger.error("Token JWT do usuário não encontrado na requisição.")
            return Response(
                {"message": "Token JWT do usuário não encontrado na requisição. Verifique sua autenticação."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        api_headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {user_jwt_token}", 
        }

        workbook_resultado = openpyxl.Workbook()
        sheet_resultado = workbook_resultado.active
        sheet_resultado.title = "Resultados da Consulta"

        # 2. Adicionar e estilizar o cabeçalho imediatamente
        headers_resultado = [col["header"] for col in self.cnpj_column_defs]
        sheet_resultado.append(headers_resultado)

        for col_num, col_def in enumerate(self.cnpj_column_defs, 1):
            cell = sheet_resultado.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = thin_border
            sheet_resultado.column_dimensions[get_column_letter(col_num)].width = (
                col_def["width"]
            )
        
        logger.info(f"Processando {len(cnpjs_from_excel)} CNPJs. Inserindo dados diretamente na planilha.")

        # 3. Processar e adicionar linhas à planilha item por item (evita lista 'resultados' grande)
        for i, item in enumerate(cnpjs_from_excel):
            cnpj_original = item.get("CNPJ")
            row_data = {} # Dicionário para formatar a linha atual

            if not cnpj_original:
                logger.warning(f"CNPJ não fornecido na linha {i+1}. Registrando erro.")
                row_data = self._format_error_result("", "CNPJ não fornecido na linha da planilha.")
            else:
                api_request_body = {
                    "tipo_consulta": "cnpj",
                    "parametro_consulta": cnpj_original,
                    "origem": origem_consulta,
                    "lote_id": str(batch_id_para_planilha) if batch_id_para_planilha else None
                }
                try:
                    # Usar um Session do requests para reutilizar conexões (pequena otimização)
                    # No entanto, a criação da sessão pode ter um custo inicial.
                    # Para muitos CNPJs, vale a pena. Para poucos, o overhead pode ser maior.
                    # Mantenha requests.post direto se preferir simplicidade ou se os lotes forem pequenos.
                    
                    # Se você decidir usar requests.Session, crie-o FORA do loop e feche-o no finally
                    
                    response_api = requests.post(
                        YOUR_CONSULTA_API_URL,
                        json=api_request_body,
                        headers=api_headers,
                        timeout=settings.API_CONSULTA_TIMEOUT,
                        verify=False# Usar uma configuração de timeout
                    )
                    response_api.raise_for_status()
                    
                    api_response_data = response_api.json()
                    
                    cnpj_data_from_api = api_response_data.get("resultado_api")

                    if (
                        api_response_data.get("mensagem") == "Consulta realizada com sucesso."
                        and cnpj_data_from_api
                    ):
                        row_data = self._format_success_result(
                            cnpj_original, cnpj_data_from_api
                        )
                        logger.debug(f"CNPJ {cnpj_original} consultado com sucesso.")
                    else:
                        error_msg = api_response_data.get(
                            "mensagem", "Resposta da API de consulta incompleta ou inesperada."
                        )
                        row_data = self._format_error_result(cnpj_original, f"API: {error_msg}")
                        logger.warning(f"CNPJ {cnpj_original} falhou: {error_msg}")

                except requests.exceptions.RequestException as e:
                    row_data = self._format_error_result(
                        cnpj_original, f"Erro na comunicação com a API de consulta: {str(e)}"
                    )
                    logger.error(f"Erro de comunicação para CNPJ {cnpj_original}: {e}")
                except Exception as e:
                    row_data = self._format_error_result(
                        cnpj_original, f"Erro inesperado no processamento: {str(e)}"
                    )
                    logger.critical(f"Erro inesperado para CNPJ {cnpj_original}: {e}", exc_info=True)

            row_values = [row_data.get(col["key"], "") for col in self.cnpj_column_defs]
            sheet_resultado.append(row_values)
           
        
        from io import BytesIO
        output_buffer = BytesIO()
        workbook_resultado.save(output_buffer)
        output_buffer.seek(0) 

        response = HttpResponse(
            output_buffer.getvalue(), 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="planilha-resultado-cnpj.xlsx"'
        )
        logger.info("Planilha de resultados gerada e pronta para download.")
        return response