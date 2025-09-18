import logging
from io import BytesIO # Importar BytesIO para manipulação de arquivos em memória
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.authentication import JWTAuthentication
import requests
import uuid
import json 
from ..serializers import ProcessamentoPlanilhaCpfInputSerializer 
from .base_views import _base_download_model_excel, _base_format_error_result
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.conf import settings # Importar settings para usar configurações do Django

logger = logging.getLogger(__name__)

# --- Recomendo fortemente que esta URL não seja fixa no código ---
# Ela DEVE vir de uma variável de ambiente em produção.
YOUR_CONSULTA_API_URL = settings.CONSULTA_API_URL



@api_view(["GET"])
@permission_classes([AllowAny]) # type: ignore
def baixar_planilha_modelo_drf_cpf(request):
    """
    View para gerar e baixar a planilha modelo de consulta de CPFs.
    """
    return _base_download_model_excel(
        sheet_title="CPFs para Consulta",
        header_name="CPF",
        example_values=["000.000.000-00", "00000000000", "123.456.789-00"],
        column_width=20,
        filename="planilha-modelo-cpf.xlsx",
    )


class ProcessarPlanilhaCpfsView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated] # Garante que o usuário esteja autenticado

    cpf_column_defs = [
        {"header": "CPF (Entrada)", "key": "CPF (Entrada)", "width": 20},
        {"header": "STATUS CONSULTA", "key": "STATUS CONSULTA", "width": 18},
        {"header": "Nome Completo", "key": "Nome Completo", "width": 40},
        {"header": "CPF", "key": "CPF", "width": 20},
        {"header": "Situação Cadastral", "key": "Situação Cadastral", "width": 25},
        {"header": "Data de Nascimento", "key": "Data de Nascimento", "width": 20},
        {"header": "Nome da Mãe", "key": "Nome da Mãe", "width": 40},
        {"header": "Nome do Pai", "key": "Nome do Pai", "width": 40},
        {"header": "Gênero", "key": "Gênero", "width": 10},
        {"header": "Idade", "key": "Idade", "width": 10},
        {"header": "Nome Comum (Alias)", "key": "Nome Comum (Alias)", "width": 30},
        {"header": "ERRO_DETALHES", "key": "ERRO_DETALHES", "width": 40},
    ]

    cpf_result_keys = [
        col["key"]
        for col in cpf_column_defs
        if col["key"] not in ["STATUS CONSULTA", "ERRO_DETALHES"]
    ]

    def _format_success_result(self, cpf_entrada, api_cpf_data):
        """
        Formata os dados de sucesso para a linha da planilha de CPF,
        extraindo-os corretamente do `resultado_api`.
        Assume que api_cpf_data é o conteúdo do 'resultado_api' da sua API principal.
        """
        basic_data = {}
        if api_cpf_data and api_cpf_data.get('Result') and isinstance(api_cpf_data['Result'], list) and len(api_cpf_data['Result']) > 0:
            basic_data = api_cpf_data['Result'][0].get('BasicData', {})
        
        aliases_data = basic_data.get('Aliases', {})
        common_name = aliases_data.get('CommonName', '')

        return {
            "CPF (Entrada)": cpf_entrada,
            "STATUS CONSULTA": "SUCESSO",
            "ERRO_DETALHES": "",
            "Nome Completo": basic_data.get("Name", ""),
            "CPF": basic_data.get("TaxIdNumber", ""),
            "Situação Cadastral": basic_data.get("TaxIdStatus", ""),
            "Data de Nascimento": basic_data.get("BirthDate", "").split("T")[0] if basic_data.get("BirthDate") else "",
            "Nome da Mãe": basic_data.get("MotherName", ""),
            "Nome do Pai": basic_data.get("FatherName", ""),
            "Gênero": basic_data.get("Gender", ""),
            "Idade": basic_data.get("Age", ""),
            "Nome Comum (Alias)": common_name,
        }

    def _format_error_result(self, cpf_entrada, error_message):
        """
        Formata os dados de erro para a linha da planilha de CPF.
        """
        return _base_format_error_result(cpf_entrada, error_message, self.cpf_result_keys)


    def post(self, request, *args, **kwargs):
        logger.info("Iniciando ProcessarPlanilhaCpfsView.post")
        if not YOUR_CONSULTA_API_URL:
            logger.error("URL da API de consulta não configurada no backend.")
            return Response(
                {"message": "URL da API de consulta não configurada no backend."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # A autenticação já ocorreu via JWTCookieAuthentication.
        # O token válido está disponível em request.auth.
        if not request.auth:
            logger.error("Token JWT não disponível para reuso na requisição interna.")
            return Response(
                {"message": "Token JWT não disponível para reuso na requisição interna. Verifique sua autenticação."},
                status=status.HTTP_401_UNAUTHORIZED,
            )
        
        # Converte o objeto de token (geralmente AccessToken) para sua representação de string
        user_jwt_token = str(request.auth) 

        serializer = ProcessamentoPlanilhaCpfInputSerializer(data=request.data)
        if not serializer.is_valid():
            logger.warning(f"Dados do serializer inválidos: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        valid_data = serializer.validated_data
        cpfs_from_excel = valid_data.get("cpfs", [])
        
        # Mantenha o valor de origem conforme o frontend envia (e.g., "planilha" ou "planilha_upload")
        origem_consulta = valid_data.get("origem", "manual") 

        batch_id_para_planilha = None
        # Ajuste a verificação 'origem_consulta' para o valor exato que o frontend envia
        # Por exemplo, se o frontend envia "planilha_upload", mude a linha abaixo
        if origem_consulta == "planilha_upload": # Ou "planilha", dependendo do seu frontend
            batch_id_para_planilha = uuid.uuid4()
            
        api_headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {user_jwt_token}", # Usa o token obtido de request.auth
        }

        # --- Otimização de Memória e Desempenho ---
        # 1. Gerar workbook e sheet imediatamente
        workbook_resultado = openpyxl.Workbook()
        sheet_resultado = workbook_resultado.active
        sheet_resultado.title = "Resultados da Consulta CPF"

        # 2. Adicionar e estilizar o cabeçalho imediatamente
        headers_resultado = [col["header"] for col in self.cpf_column_defs]
        sheet_resultado.append(headers_resultado)

        for col_num, col_def in enumerate(self.cpf_column_defs, 1):
            cell = sheet_resultado.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = thin_border
            sheet_resultado.column_dimensions[get_column_letter(col_num)].width = (
                col_def["width"]
            )
        
        logger.info(f"Processando {len(cpfs_from_excel)} CPFs. Inserindo dados diretamente na planilha.")

        # 3. Processar e adicionar linhas à planilha item por item (evita lista 'resultados' grande)
        for i, item in enumerate(cpfs_from_excel):
            cpf_original = item.get("CPF")
            row_data = {} # Dicionário para formatar a linha atual

            if not cpf_original:
                logger.warning(f"CPF não fornecido na linha {i+1} da planilha. Registrando erro.")
                row_data = self._format_error_result("", "CPF não fornecido na linha da planilha.")
            else:
                tipo_consulta = "cpf" 
                parametro_para_api = cpf_original

                api_request_body = {
                    "tipo_consulta": tipo_consulta,
                    "parametro_consulta": parametro_para_api,
                    "origem": origem_consulta,
                    "lote_id": str(batch_id_para_planilha) if batch_id_para_planilha else None
                }
                
                try:
                    response_api = requests.post(
                        YOUR_CONSULTA_API_URL,
                        json=api_request_body,
                        headers=api_headers,
                        timeout=settings.API_CONSULTA_TIMEOUT,
                        verify=False # Usar uma configuração de timeout
                    )
                    response_api.raise_for_status() # Levanta um erro para status 4xx/5xx
                    api_response_data = response_api.json()

                    cpf_data_from_api = api_response_data.get("resultado_api")

                    if api_response_data.get("mensagem") == "Consulta realizada com sucesso." and cpf_data_from_api:
                        row_data = self._format_success_result(cpf_original, cpf_data_from_api)
                        logger.debug(f"CPF {cpf_original} consultado com sucesso.")
                    else:
                        error_msg = api_response_data.get(
                            "mensagem", "Resposta da API incompleta ou inesperada."
                        )
                        row_data = self._format_error_result(cpf_original, f"API: {error_msg}")
                        logger.warning(f"CPF {cpf_original} falhou: {error_msg}")

                except requests.exceptions.RequestException as e:
                    status_code = e.response.status_code if e.response is not None else 'N/A'
                    response_content = e.response.text if e.response is not None else 'Sem conteúdo de resposta.'
                    row_data = self._format_error_result(
                        cpf_original, f"Erro na consulta à API ({status_code}): {str(e)}. Detalhes: {response_content}"
                    )
                    logger.error(f"Erro de comunicação para CPF {cpf_original}: {e}. Status: {status_code}. Conteúdo: {response_content}")
                except Exception as e:
                    row_data = self._format_error_result(
                        cpf_original, f"Erro inesperado no processamento: {str(e)}"
                    )
                    logger.critical(f"Erro inesperado para CPF {cpf_original}: {e}", exc_info=True)

            row_values = [row_data.get(col["key"], "") for col in self.cpf_column_defs]
            sheet_resultado.append(row_values)

        output_buffer = BytesIO()
        workbook_resultado.save(output_buffer)
        output_buffer.seek(0) 

        response = HttpResponse(
            output_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="planilha-resultado-cpf.xlsx"'
        )
        logger.info("Planilha de resultados de CPF gerada e pronta para download.")
        return response