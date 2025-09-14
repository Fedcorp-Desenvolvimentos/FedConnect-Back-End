import logging
from io import BytesIO

from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny



from rest_framework_simplejwt.authentication import JWTAuthentication 

import requests
import uuid

# Você precisará criar ProcessamentoPlanilhaCepInputSerializer
from ..serializers import ProcessamentoPlanilhaCepInputSerializer 
from .base_views import _base_download_model_excel, _base_format_error_result

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.conf import settings # Importe settings para usar configurações do Django

logger = logging.getLogger(__name__)

YOUR_CONSULTA_API_URL = settings.CONSULTA_API_URL


@api_view(["GET"])
@permission_classes([AllowAny]) # type: ignore
def baixar_planilha_modelo_drf_cep(request):
    """
    View para gerar e baixar a planilha modelo de consulta de CEPs.
    """
    return _base_download_model_excel(
        sheet_title="CEPs para Consulta",
        header_name="CEP",
        example_values=["20000-000", "22000000", "99999-999"],
        column_width=20,
        filename="planilha-modelo-cep.xlsx",
    )


class ProcessarPlanilhaCepsView(APIView):
    
    authentication_classes = [JWTAuthentication ] 
    permission_classes = [IsAuthenticated] # Garante que apenas usuários autenticados possam acessar esta view

    cep_column_defs = [
        {"header": "CEP (Entrada)", "key": "CEP (Entrada)", "width": 15},
        {"header": "STATUS CONSULTA", "key": "STATUS CONSULTA", "width": 18},
        {"header": "CEP", "key": "CEP", "width": 15},
        {"header": "Logradouro", "key": "Logradouro", "width": 40},
        {"header": "Bairro", "key": "Bairro", "width": 25},
        {"header": "Cidade", "key": "Cidade", "width": 25},
        {"header": "UF", "key": "UF", "width": 10},
        {"header": "Complemento", "key": "Complemento", "width": 20},
        {"header": "ERRO_DETALHES", "key": "ERRO_DETALHES", "width": 40},
    ]

    cep_result_keys = [
        col["key"]
        for col in cep_column_defs
        if col["key"] not in ["STATUS CONSULTA", "ERRO_DETALHES"]
    ]

    def _format_success_result(self, cep_entrada, api_cep_data):
        """
        Formata os dados de sucesso para a linha da planilha de CEP.
        Ajuste os 'get's para corresponder aos nomes dos campos retornados pela sua API de CEP.
        """
        return {
            "CEP (Entrada)": cep_entrada,
            "STATUS CONSULTA": "SUCESSO",
            "ERRO_DETALHES": "",
            "CEP": api_cep_data.get("cep", ""),
            "Logradouro": api_cep_data.get("street", ""), # <-- Verifique o nome real do campo
            "Bairro": api_cep_data.get("neighborhood", ""), # <-- Verifique o nome real do campo
            "Cidade": api_cep_data.get("city", ""), # <-- 'localidade' é comum, verifique sua API
            "UF": api_cep_data.get("state", ""),
            "Complemento": api_cep_data.get("complement", ""),
        }

    def _format_error_result(self, cep_entrada, error_message):
        """
        Formata os dados de erro para a linha da planilha de CEP, utilizando o helper base.
        """
        return _base_format_error_result(cep_entrada, error_message, self.cep_result_keys)


    def post(self, request, *args, **kwargs):
        logger.info("Iniciando ProcessarPlanilhaCepsView.post")
        if not YOUR_CONSULTA_API_URL:
            logger.error("URL da API de consulta não configurada no backend.")
            return Response(
                {"message": "URL da API de consulta não configurada no backend."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        user_jwt_token = None
        if request.auth:
            user_jwt_token = str(request.auth)

        if not user_jwt_token:
            logger.error("Token JWT do usuário não encontrado na requisição.")
            return Response(
                {"message": "Token JWT do usuário não encontrado na requisição. Verifique sua autenticação."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        serializer = ProcessamentoPlanilhaCepInputSerializer(data=request.data)
        if not serializer.is_valid():
            logger.warning(f"Dados do serializer inválidos: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        valid_data = serializer.validated_data
        ceps_from_excel = valid_data.get("ceps", []) 
        origem_consulta = valid_data.get("origem", "planilha")

        batch_id_para_planilha = None
        if origem_consulta == "planilha":
            batch_id_para_planilha = uuid.uuid4()
            
        api_headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {user_jwt_token}", 
        }
        
        # --- Otimização de Memória e Desempenho ---
        # 1. Gerar workbook e sheet imediatamente
        workbook_resultado = openpyxl.Workbook()
        sheet_resultado = workbook_resultado.active
        sheet_resultado.title = "Resultados da Consulta CEP"

        # 2. Adicionar e estilizar o cabeçalho imediatamente
        headers_resultado = [col["header"] for col in self.cep_column_defs] # Use cep_column_defs
        sheet_resultado.append(headers_resultado)

        for col_num, col_def in enumerate(self.cep_column_defs, 1):
            cell = sheet_resultado.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = thin_border
            sheet_resultado.column_dimensions[get_column_letter(col_num)].width = (
                col_def["width"]
            )
        
        logger.info(f"Processando {len(ceps_from_excel)} CEPs. Inserindo dados diretamente na planilha.")

        # 3. Processar e adicionar linhas à planilha item por item (evita lista 'resultados' grande)
        for i, item in enumerate(ceps_from_excel): 
            cep_original = item.get("CEP")
            row_data = {} # Dicionário para formatar a linha atual

            if not cep_original:
                logger.warning(f"CEP não fornecido na linha {i+1} da planilha. Registrando erro.")
                row_data = self._format_error_result("", "CEP não fornecido na linha da planilha.")
            else:
                api_request_body = {
                    "tipo_consulta": "endereco", # Ajuste para "endereco" ou o que você usa para consultas de CEP
                    "parametro_consulta": cep_original,
                    "origem": origem_consulta,
                    "lote_id": str(batch_id_para_planilha) if batch_id_para_planilha else None
                }
                
                try:
                    response_api = requests.post(
                        YOUR_CONSULTA_API_URL,
                        json=api_request_body,
                        headers=api_headers,
                        timeout=settings.API_CONSULTA_TIMEOUT, # Usar uma configuração de timeout
                    )
                    response_api.raise_for_status() 
                    api_response_data = response_api.json()
                    
                    cep_data_from_api = api_response_data.get("resultado_api")

                    if (
                        api_response_data.get("mensagem") == "Consulta realizada com sucesso."
                        and cep_data_from_api
                    ):
                        row_data = self._format_success_result(
                            cep_original, cep_data_from_api
                        )
                        logger.debug(f"CEP {cep_original} consultado com sucesso.")
                    else:
                        error_msg = api_response_data.get(
                            "mensagem", "Resposta da API de consulta de CEP incompleta ou inesperada."
                        )
                        row_data = self._format_error_result(cep_original, f"API: {error_msg}")
                        logger.warning(f"CEP {cep_original} falhou: {error_msg}")

                except requests.exceptions.RequestException as e:
                    row_data = self._format_error_result(
                        cep_original, f"Erro na comunicação com a API de consulta: {str(e)}"
                    )
                    logger.error(f"Erro de comunicação para CEP {cep_original}: {e}")
                except Exception as e:
                    row_data = self._format_error_result(
                        cep_original, f"Erro inesperado no processamento: {str(e)}"
                    )
                    logger.critical(f"Erro inesperado para CEP {cep_original}: {e}", exc_info=True)

            # Adiciona a linha formatada diretamente à planilha
            row_values = [row_data.get(col["key"], "") for col in self.cep_column_defs]
            sheet_resultado.append(row_values)
            
        # Configura a resposta HTTP para download do arquivo Excel
        # Usamos BytesIO para salvar o workbook na memória e evitar escrever em disco
        output_buffer = BytesIO()
        workbook_resultado.save(output_buffer)
        output_buffer.seek(0) # Volta ao início do buffer

        response = HttpResponse(
            output_buffer.getvalue(), # Obtém os bytes do buffer
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="planilha-resultado-cep.xlsx"'
        )
        logger.info("Planilha de resultados de CEP gerada e pronta para download.")
        return response