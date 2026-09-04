# condomed/views.py
import openpyxl
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Q
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from users.permissions import IsCondomedOrAdmin

from . import services
from .models import INSTRUTORES_CIPA, LOCAIS_CIPA, InscricaoCipa, TurmaCipa
from .serializers import (
    CPF_DUPLICADO,
    ImportarTurmaSerializer,
    InscricaoCipaSerializer,
    InscricaoComTurmaSerializer,
    TurmaCipaSerializer,
    TurmaResumoSerializer,
)
from .validators import cpf_valido, normalizar_cpf


class PaginacaoHistorico(PageNumberPagination):
    """Paginação das consultas de histórico (RF-HIS-001, RF-HIS-002).

    O calendário continua sem paginação (pede o mês inteiro); só as listas
    abertas por período paginam. 25 por página cabe numa tela sem rolagem
    infinita; `page_size` deixa a tela pedir mais quando fizer sentido.
    """

    page_size = 25
    page_size_query_param = "page_size"
    max_page_size = 100

# Colunas da planilha modelo, na ordem. O cabeçalho é o contrato com a tela:
# o parser do frontend casa por este texto, então mudar aqui é mudar lá.
def salvar_inscricao(serializer, **kwargs):
    """Grava a inscrição traduzindo a constraint de CPF duplicado em 400.

    O `unique_together (turma, cpf)` é a garantia final da regra — a validação
    do serializer não cobre duas requisições simultâneas com o mesmo CPF, que
    passam as duas pela checagem e só se encontram no banco. Sem esta
    tradução, a segunda viraria 500.

    É a única constraint de unicidade da tabela, então IntegrityError aqui só
    pode ser ela. O `atomic` isola a falha: no PostgreSQL, a transação fica
    inutilizável depois de um IntegrityError.
    """
    try:
        with transaction.atomic():
            return serializer.save(**kwargs)
    except IntegrityError as erro:
        raise ValidationError({"cpf": CPF_DUPLICADO}) from erro


COLUNAS_MODELO = [
    ("administradora", 28, "Delforte Administração"),
    ("condominio", 28, "Residencial Aurora"),
    ("cnpj_condominio", 20, "01.998.690/0001-82"),
    ("nome", 30, "Maria Aparecida da Silva"),
    ("cpf", 16, "529.982.247-25"),
    ("funcao", 20, "Zeladora"),
    ("email", 26, "maria@exemplo.com.br"),
    ("telefone", 18, "11999998888"),
]


class TurmaCipaViewSet(viewsets.ModelViewSet):
    """Turmas do curso CIPA (RF-CIP-001..004). Acesso: condomed + admin."""

    serializer_class = TurmaCipaSerializer
    permission_classes = [IsCondomedOrAdmin]
    # O calendário pede o mês inteiro de uma vez; sem paginação a resposta é uma lista.
    pagination_class = None
    queryset = TurmaCipa.objects.select_related("reserva_sala").prefetch_related(
        "inscricoes"
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        local = params.get("local")
        mes = params.get("mes")
        ano = params.get("ano")

        if local:
            queryset = queryset.filter(local=local)
        if ano:
            queryset = queryset.filter(data__year=ano)
        if mes:
            queryset = queryset.filter(data__month=mes)
        return queryset

    @transaction.atomic
    def perform_create(self, serializer):
        # Turma + Reserva espelho na mesma transação (RNF-CIP-002).
        turma = serializer.save(criado_por=self.request.user)
        services.sincronizar_espelho(turma, self.request.user)

    @transaction.atomic
    def perform_update(self, serializer):
        turma = serializer.save()
        services.sincronizar_espelho(turma, self.request.user)

    @transaction.atomic
    def perform_destroy(self, instance):
        services.remover_espelho(instance)
        instance.delete()

    @action(detail=False, methods=["get"], url_path="locais")
    def locais(self, request):
        """Locais e capacidades para montar as abas do frontend."""
        return Response([
            {"codigo": codigo, **dados} for codigo, dados in LOCAIS_CIPA.items()
        ])

    @action(detail=False, methods=["get"], url_path="historico")
    def historico(self, request):
        """Turmas por período, paginadas, com filtros que o calendário não tem.

        Rota separada de `GET cursos-cipa/` de propósito: o calendário depende
        daquela devolver o mês inteiro como lista, sem envelope de paginação.
        """
        params = request.query_params
        queryset = (
            TurmaCipa.objects.select_related("reserva_sala")
            .prefetch_related("inscricoes")
            .order_by("-data", "local")
        )

        if params.get("data_inicio"):
            queryset = queryset.filter(data__gte=params["data_inicio"])
        if params.get("data_fim"):
            queryset = queryset.filter(data__lte=params["data_fim"])
        if params.get("local"):
            queryset = queryset.filter(local=params["local"])
        if params.get("status"):
            queryset = queryset.filter(status=params["status"])
        # Vínculo é do inscrito (ADR-0004): filtrar turma por administradora ou
        # condomínio é "turmas que têm gente de lá".
        if params.get("administradora"):
            queryset = queryset.filter(
                inscricoes__administradora_codigo=params["administradora"]
            )
        if params.get("condominio"):
            queryset = queryset.filter(
                inscricoes__condominio_nome__icontains=params["condominio"]
            )
        if params.get("busca"):
            termo = params["busca"].strip()
            queryset = queryset.filter(
                Q(inscricoes__nome__icontains=termo)
                | Q(inscricoes__cpf=normalizar_cpf(termo))
                | Q(inscricoes__condominio_nome__icontains=termo)
                | Q(inscricoes__administradora_nome__icontains=termo)
                | Q(observacao__icontains=termo)
            )
        queryset = queryset.distinct()

        paginador = PaginacaoHistorico()
        pagina = paginador.paginate_queryset(queryset, request, view=self)
        return paginador.get_paginated_response(
            TurmaResumoSerializer(pagina, many=True).data
        )

    @action(detail=False, methods=["get"], url_path="participantes")
    def participantes(self, request):
        """Inscrições em todas as turmas: onde cada pessoa esteve.

        Generaliza o `verificar-cpf`, que segue existindo para a tela de
        inscrição. Uma linha por inscrição — a mesma pessoa em três turmas são
        três linhas, porque presença e certificado (fases seguintes) são por
        inscrição.
        """
        params = request.query_params
        queryset = InscricaoCipa.objects.select_related("turma").order_by(
            "-turma__data", "condominio_nome", "nome"
        )

        if params.get("cpf"):
            queryset = queryset.filter(cpf=normalizar_cpf(params["cpf"]))
        if params.get("administradora"):
            queryset = queryset.filter(administradora_codigo=params["administradora"])
        if params.get("condominio"):
            queryset = queryset.filter(condominio_nome__icontains=params["condominio"])
        if params.get("data_inicio"):
            queryset = queryset.filter(turma__data__gte=params["data_inicio"])
        if params.get("data_fim"):
            queryset = queryset.filter(turma__data__lte=params["data_fim"])
        if params.get("busca"):
            termo = params["busca"].strip()
            digitos = normalizar_cpf(termo)
            filtro = (
                Q(nome__icontains=termo)
                | Q(condominio_nome__icontains=termo)
                | Q(administradora_nome__icontains=termo)
            )
            # Só compara CPF quando o termo tem dígitos: "Maria" não é CPF.
            if digitos:
                filtro |= Q(cpf__startswith=digitos)
            queryset = queryset.filter(filtro)

        paginador = PaginacaoHistorico()
        pagina = paginador.paginate_queryset(queryset, request, view=self)
        return paginador.get_paginated_response(
            InscricaoComTurmaSerializer(pagina, many=True).data
        )

    @action(detail=False, methods=["get"], url_path="planilha-modelo")
    def planilha_modelo(self, request):
        """Planilha modelo dos inscritos, com uma linha de exemplo.

        Local e data são escolhidos na tela, não na planilha: uma planilha é
        uma turma, e data errada em célula é erro caro de achar.
        """
        workbook = openpyxl.Workbook()
        aba = workbook.active
        aba.title = "Inscritos"

        for indice, (cabecalho, largura, exemplo) in enumerate(COLUNAS_MODELO, 1):
            celula = aba.cell(row=1, column=indice, value=cabecalho)
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill("solid", start_color="0F3D5D")
            celula.alignment = Alignment(horizontal="center")
            aba.column_dimensions[get_column_letter(indice)].width = largura
            aba.cell(row=2, column=indice, value=exemplo)

        aba.freeze_panes = "A2"
        # CPF e CNPJ ficam como texto: formatados como número, o Excel come o
        # zero à esquerda e a linha volta inválida.
        colunas_texto = [
            indice
            for indice, (cabecalho, _, _) in enumerate(COLUNAS_MODELO, 1)
            if cabecalho in ("cpf", "cnpj_condominio")
        ]
        for linha in range(2, 200):
            for coluna in colunas_texto:
                aba.cell(row=linha, column=coluna).number_format = "@"

        resposta = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        resposta["Content-Disposition"] = (
            'attachment; filename="modelo-inscritos-cipa.xlsx"'
        )
        workbook.save(resposta)
        return resposta

    @action(detail=False, methods=["post"], url_path="importar")
    def importar(self, request):
        """Cria a turma com os inscritos da planilha, tudo na mesma transação.

        Ou nasce completa, ou não nasce: turma vazia por falha na segunda
        metade seria pior que erro nenhum. A tela já valida linha a linha e só
        envia o que passou, então aqui o payload é tudo-ou-nada.
        """
        serializer = ImportarTurmaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            turma = serializer.criar(request.user)

        return Response(
            TurmaCipaSerializer(turma).data, status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=["get"], url_path="instrutores")
    def instrutores(self, request):
        """Instrutores que assinam o certificado, para o select da turma.

        Lista fixa no código (decisão do dono, 2026-09-04): sem cadastro
        editável não há como cadastrar errado. A assinatura fica só no
        servidor — o nome do arquivo não sai daqui.
        """
        return Response([
            {
                "codigo": codigo,
                "nome": dados["nome"],
                "titulo": dados["titulo"],
                "registro": f"MTE/{dados['registro_uf']} {dados['registro_mte']}",
            }
            for codigo, dados in INSTRUTORES_CIPA.items()
        ])

    @action(detail=False, methods=["get"], url_path="verificar-cpf")
    def verificar_cpf(self, request):
        """Onde mais este CPF já está inscrito.

        Duplicidade na mesma turma é barrada no serializer; entre turmas é
        permitida, e a tela usa esta consulta para avisar antes de gravar.
        `excluir_turma` tira da resposta a turma que está sendo preenchida.
        """
        cpf = normalizar_cpf(request.query_params.get("cpf", ""))
        if not cpf_valido(cpf):
            return Response(
                {"detail": "CPF inválido."}, status=status.HTTP_400_BAD_REQUEST
            )

        inscricoes = (
            InscricaoCipa.objects.filter(cpf=cpf)
            .select_related("turma")
            .order_by("-turma__data")
        )
        excluir = request.query_params.get("excluir_turma")
        if excluir:
            inscricoes = inscricoes.exclude(turma_id=excluir)

        return Response([
            {
                "inscricao_id": inscricao.id,
                "turma_id": inscricao.turma_id,
                "nome": inscricao.nome,
                "data": inscricao.turma.data,
                "local": inscricao.turma.local,
                "local_nome": LOCAIS_CIPA.get(inscricao.turma.local, {}).get(
                    "nome", inscricao.turma.local
                ),
                # O vínculo é da inscrição (ADR-0004): é o que distingue
                # "repetiu a pessoa" de "a mesma pessoa vem por dois condomínios".
                "administradora_codigo": inscricao.administradora_codigo,
                "administradora_nome": inscricao.administradora_nome,
                "condominio_nome": inscricao.condominio_nome,
                "status": inscricao.turma.status,
            }
            for inscricao in inscricoes
        ])

    @action(detail=True, methods=["get", "post"], url_path="inscricoes")
    def inscricoes(self, request, pk=None):
        if request.method == "GET":
            turma = self.get_object()
            serializer = InscricaoCipaSerializer(
                turma.inscricoes.all(), many=True
            )
            return Response(serializer.data)

        # Sem `select_for_update`: ele existia só para a capacidade não ser
        # estourada em corrida (INV-CIP-003), e a capacidade deixou de ser
        # limite (ADR-0006). A unicidade de CPF por turma segue no
        # `unique_together`, e `salvar_inscricao` traduz a constraint em 400.
        turma = self.get_object()
        serializer = InscricaoCipaSerializer(
            data=request.data, context={"turma": turma, "request": request}
        )
        serializer.is_valid(raise_exception=True)
        salvar_inscricao(serializer, turma=turma)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(
        detail=True,
        methods=["patch", "delete"],
        url_path=r"inscricoes/(?P<inscricao_id>\d+)",
    )
    def inscricao_detalhe(self, request, pk=None, inscricao_id=None):
        turma = self.get_object()
        inscricao = turma.inscricoes.filter(pk=inscricao_id).first()
        if inscricao is None:
            return Response(
                {"detail": "Inscrição não encontrada nesta turma."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if request.method == "DELETE":
            inscricao.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        serializer = InscricaoCipaSerializer(
            inscricao, data=request.data, partial=True, context={"turma": turma}
        )
        serializer.is_valid(raise_exception=True)
        salvar_inscricao(serializer)
        return Response(serializer.data)
